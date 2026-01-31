"""
Case Evaluator
==============
Author: AeroGuardian Member
Date: 2026-01-21

Unified evaluator that orchestrates all 4 research-grade metrics:
- SFS: Scenario Fidelity Score
- BRR: Behavior Reproduction Rate
- ECC: Evidence-Conclusion Consistency
- ESRI: Executable Safety Reliability Index

Produces per-incident and aggregate evaluation reports.
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from dataclasses import dataclass

logger = logging.getLogger("AeroGuardian.Evaluator.Case")

# Import metric components
from .scenario_fidelity import ScenarioFidelityScorer
from .behavior_validation import BehaviorValidator
from .evidence_consistency import EvidenceConsistencyChecker
from .esri import ESRICalculator


@dataclass
class CaseEvaluationResult:
    """Complete evaluation result for a single incident case."""
    
    incident_id: str
    evaluation_timestamp: str
    
    # Metric results
    sfs: float
    brr: float
    ecc: float
    esri: float
    
    # Trust assessment
    trust_level: str
    trust_justification: str
    
    # Detailed breakdowns
    sfs_details: Dict
    brr_details: Dict
    ecc_details: Dict
    
    # Anomaly summary
    detected_anomalies: List[Dict]
    unsupported_claims: List[str]
    
    def to_dict(self) -> Dict:
        return {
            "incident_id": self.incident_id,
            "evaluation_timestamp": self.evaluation_timestamp,
            "scores": {
                "ESRI": round(self.esri, 4),
                "SFS": round(self.sfs, 3),
                "BRR": round(self.brr, 3),
                "ECC": round(self.ecc, 3),
            },
            "trust_level": self.trust_level,
            "trust_justification": self.trust_justification,
            "details": {
                "scenario_fidelity": self.sfs_details,
                "behavior_reproduction": self.brr_details,
                "evidence_consistency": self.ecc_details,
            },
            "detected_anomalies": self.detected_anomalies,
            "unsupported_claims": self.unsupported_claims,
        }
    
    def to_row(self) -> Dict:
        """Convert to flat row for Excel export."""
        return {
            "Incident ID": self.incident_id,
            "Timestamp": self.evaluation_timestamp,
            "ESRI": round(self.esri, 4),
            "SFS": round(self.sfs, 3),
            "BRR": round(self.brr, 3),
            "ECC": round(self.ecc, 3),
            "Trust Level": self.trust_level,
            "Anomaly Count": len(self.detected_anomalies),
            "Unsupported Claims": len(self.unsupported_claims),
            "SFS Confidence": self.sfs_details.get("confidence", ""),
            "BRR Confidence": self.brr_details.get("confidence", ""),
            "ECC Confidence": self.ecc_details.get("confidence", ""),
        }


class CaseEvaluator:
    """
    Orchestrates evaluation of a single incident case using all 4 metrics.
    
    Usage:
        evaluator = CaseEvaluator()
        result = evaluator.evaluate(
            faa_report=incident,
            px4_config=config,
            telemetry=telemetry_data,
            safety_report=safety_analysis,
        )
    """
    
    def __init__(self):
        self.sfs_scorer = ScenarioFidelityScorer()
        self.brr_validator = BehaviorValidator()
        self.ecc_checker = EvidenceConsistencyChecker()
        self.esri_calculator = ESRICalculator()
        
        logger.debug("CaseEvaluator initialized with all metric components")
    
    def evaluate(
        self,
        faa_report: Dict,
        px4_config: Dict,
        telemetry: List[Dict],
        safety_report: Dict,
        telemetry_stats: Optional[Dict] = None,
    ) -> CaseEvaluationResult:
        """
        Evaluate a complete incident case.
        
        Args:
            faa_report: Original FAA incident data
            px4_config: LLM-generated PX4 configuration
            telemetry: Raw telemetry data points
            safety_report: Generated safety report
            telemetry_stats: Pre-computed telemetry statistics (optional)
            
        Returns:
            CaseEvaluationResult with all metrics and trust assessment
        """
        incident_id = faa_report.get("incident_id", px4_config.get("faa_source", {}).get("incident_id", "unknown"))
        
        logger.info(f"Evaluating case: {incident_id}")
        
        # 1. Compute SFS (Scenario Fidelity Score)
        sfs_result = self.sfs_scorer.evaluate(faa_report, px4_config)
        
        # 2. Compute BRR (Behavior Reproduction Rate)
        fault_type = px4_config.get("fault_injection", {}).get("fault_type", "")
        brr_result = self.brr_validator.evaluate(telemetry, fault_type, telemetry_stats)
        
        # 3. Compute ECC (Evidence-Conclusion Consistency)
        detected_anomalies_dicts = [a.to_dict() for a in brr_result.detected_anomalies]
        stats = telemetry_stats or self.brr_validator._compute_telemetry_stats(telemetry)
        ecc_result = self.ecc_checker.evaluate(
            safety_report,
            detected_anomalies_dicts,
            stats
        )
        
        # 4. Compute ESRI (Executable Safety Reliability Index)
        esri_result = self.esri_calculator.calculate(
            sfs_result.to_dict(),
            brr_result.to_dict(),
            ecc_result.to_dict(),
            incident_id
        )
        
        # Build result
        result = CaseEvaluationResult(
            incident_id=incident_id,
            evaluation_timestamp=datetime.now().isoformat(),
            sfs=sfs_result.score,
            brr=brr_result.score,
            ecc=ecc_result.score,
            esri=esri_result.esri,
            trust_level=esri_result.trust_level,
            trust_justification=esri_result.trust_justification,
            sfs_details=sfs_result.to_dict(),
            brr_details=brr_result.to_dict(),
            ecc_details=ecc_result.to_dict(),
            detected_anomalies=detected_anomalies_dicts,
            unsupported_claims=ecc_result.unsupported_claims,
        )
        
        logger.info(
            f"Case evaluation complete: ESRI={result.esri * 100:.1f}% "
            f"(SFS={result.sfs * 100:.0f}%, BRR={result.brr * 100:.0f}%, ECC={result.ecc * 100:.0f}%) "
            f"[{result.trust_level}]"
        )
        
        return result
    
    def export_to_json(self, result: CaseEvaluationResult, output_path: Path) -> Path:
        """Export evaluation result to JSON file."""
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result.to_dict(), f, indent=2)
        
        logger.info(f"Evaluation exported to: {output_path}")
        return output_path


class EvaluationExcelExporter:
    """
    Exports evaluation results to Excel format.
    
    Generates both per-incident and aggregate reports.
    """
    
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_incident(
        self,
        result: CaseEvaluationResult,
        output_path: Optional[Path] = None
    ) -> Path:
        """Export single incident evaluation to Excel."""
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel export")
            return None
        
        output_path = output_path or (self.output_dir / f"evaluation_{result.incident_id}.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluation"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Trust level colors
        trust_fills = {
            "HIGH": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
            "LOW": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
        }
        
        # Header
        ws.merge_cells("A1:D1")
        ws["A1"] = f"AeroGuardian Evaluation Report"
        ws["A1"].font = Font(bold=True, size=14)
        
        ws.merge_cells("A2:D2")
        ws["A2"] = f"Incident: {result.incident_id}"
        
        # Scores section
        row = 4
        headers = ["Metric", "Score", "Confidence", "Details"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # ESRI row
        row += 1
        ws.cell(row=row, column=1, value="ESRI (Final Score)")
        ws.cell(row=row, column=2, value=round(result.esri, 4))
        ws.cell(row=row, column=3, value=result.trust_level)
        ws.cell(row=row, column=3).fill = trust_fills.get(result.trust_level, trust_fills["LOW"])
        ws.cell(row=row, column=4, value=result.trust_justification[:80] + "...")
        
        # SFS row
        row += 1
        ws.cell(row=row, column=1, value="SFS (Scenario Fidelity)")
        ws.cell(row=row, column=2, value=round(result.sfs, 3))
        ws.cell(row=row, column=3, value=result.sfs_details.get("confidence", ""))
        ws.cell(row=row, column=4, value=result.sfs_details.get("matched_fault_type", ""))
        
        # BRR row
        row += 1
        ws.cell(row=row, column=1, value="BRR (Behavior Reproduction)")
        ws.cell(row=row, column=2, value=round(result.brr, 3))
        ws.cell(row=row, column=3, value=result.brr_details.get("confidence", ""))
        ws.cell(row=row, column=4, value=f"{len(result.detected_anomalies)} anomalies detected")
        
        # ECC row
        row += 1
        ws.cell(row=row, column=1, value="ECC (Evidence Consistency)")
        ws.cell(row=row, column=2, value=round(result.ecc, 3))
        ws.cell(row=row, column=3, value=result.ecc_details.get("confidence", ""))
        ws.cell(row=row, column=4, value=f"{len(result.unsupported_claims)} unsupported claims")
        
        # Anomalies section
        row += 2
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Detected Anomalies"
        ws[f"A{row}"].font = Font(bold=True)
        
        row += 1
        for col, header in enumerate(["Type", "Severity", "Value", "Description"], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for anomaly in result.detected_anomalies[:10]:
            row += 1
            ws.cell(row=row, column=1, value=anomaly.get("type", ""))
            ws.cell(row=row, column=2, value=anomaly.get("severity", ""))
            ws.cell(row=row, column=3, value=round(anomaly.get("measured", 0), 2))
            ws.cell(row=row, column=4, value=anomaly.get("description", "")[:60])
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 60
        
        wb.save(output_path)
        logger.info(f"Evaluation Excel exported: {output_path}")
        return output_path
    
    def export_aggregate(
        self,
        results: List[CaseEvaluationResult],
        output_path: Optional[Path] = None
    ) -> Path:
        """Export aggregate evaluation across multiple incidents."""
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel export")
            return None
        
        output_path = output_path or (self.output_dir / "aggregate_evaluation.xlsx")
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary.merge_cells("A1:G1")
        ws_summary["A1"] = "AeroGuardian Aggregate Evaluation Report"
        ws_summary["A1"].font = Font(bold=True, size=14)
        
        ws_summary["A3"] = f"Total Incidents: {len(results)}"
        ws_summary["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Aggregate stats
        if results:
            esri_scores = [r.esri for r in results]
            ws_summary["A6"] = "Aggregate Statistics"
            ws_summary["A6"].font = Font(bold=True)
            ws_summary["A7"] = f"Average ESRI: {sum(esri_scores)/len(esri_scores):.4f}"
            ws_summary["A8"] = f"Min ESRI: {min(esri_scores):.4f}"
            ws_summary["A9"] = f"Max ESRI: {max(esri_scores):.4f}"
            
            trust_counts = {
                "HIGH": sum(1 for r in results if r.trust_level == "HIGH"),
                "MEDIUM": sum(1 for r in results if r.trust_level == "MEDIUM"),
                "LOW": sum(1 for r in results if r.trust_level == "LOW"),
            }
            ws_summary["A11"] = "Trust Distribution"
            ws_summary["A11"].font = Font(bold=True)
            ws_summary["A12"] = f"HIGH: {trust_counts['HIGH']} ({100*trust_counts['HIGH']/len(results):.1f}%)"
            ws_summary["A13"] = f"MEDIUM: {trust_counts['MEDIUM']} ({100*trust_counts['MEDIUM']/len(results):.1f}%)"
            ws_summary["A14"] = f"LOW: {trust_counts['LOW']} ({100*trust_counts['LOW']/len(results):.1f}%)"
        
        # Sheet 2: Per-Incident Details
        ws_details = wb.create_sheet("Per-Incident Details")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        headers = list(results[0].to_row().keys()) if results else []
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, result in enumerate(results, 2):
            row_data = result.to_row()
            for col_idx, header in enumerate(headers, 1):
                ws_details.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
        
        # Adjust column widths
        for col_idx, header in enumerate(headers, 1):
            ws_details.column_dimensions[get_column_letter(col_idx)].width = max(len(str(header)) + 2, 12)
        
        wb.save(output_path)
        logger.info(f"Aggregate evaluation Excel exported: {output_path}")
        return output_path


# =============================================================================
# SINGLETON ACCESSOR
# =============================================================================

_evaluator: Optional[CaseEvaluator] = None

def get_case_evaluator() -> CaseEvaluator:
    """Get singleton case evaluator instance."""
    global _evaluator
    if _evaluator is None:
        _evaluator = CaseEvaluator()
    return _evaluator


def get_column_letter(col_idx: int) -> str:
    """Convert column index to Excel letter (1=A, 2=B, etc.)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result
